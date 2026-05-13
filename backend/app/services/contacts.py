import csv
import io
from collections.abc import Iterable
from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Contact, ConsentStatus
from app.schemas import ContactCreate, ImportResult


BLOCKED_STATUSES = {ConsentStatus.unsubscribed, ConsentStatus.bounced, ConsentStatus.complained}


def normalize_email(email: str) -> str:
    return validate_email(email, check_deliverability=False).normalized.lower()


def is_valid_email(email: str) -> bool:
    try:
        normalize_email(email)
        return True
    except EmailNotValidError:
        return False


def upsert_contact(db: Session, user_id: int, payload: ContactCreate) -> tuple[Contact, bool]:
    email = normalize_email(str(payload.email))
    contact = db.scalar(select(Contact).where(Contact.user_id == user_id, Contact.email == email))
    created = False
    if contact is None:
        contact = Contact(user_id=user_id, email=email)
        created = True
        db.add(contact)

    for field, value in payload.model_dump(exclude={"email"}).items():
        setattr(contact, field, value)
    contact.email = email
    db.commit()
    db.refresh(contact)
    return contact, created


def parse_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames or "email" not in [name.lower() for name in reader.fieldnames]:
        raise ValueError("CSV must include an email column")
    return list(reader)


def parse_xlsx_bytes(data: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    if "email" not in headers:
        raise ValueError("Excel file must include an email column")
    records = []
    for row in rows[1:]:
        if not any(row):
            continue
        records.append({headers[index]: row[index] for index in range(len(headers)) if headers[index]})
    return records


def import_contacts(db: Session, user_id: int, rows: Iterable[dict], source: str) -> ImportResult:
    imported = updated = skipped = 0
    errors: list[dict] = []
    seen: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        raw_email = str(row.get("email") or "").strip()
        if not raw_email:
            skipped += 1
            errors.append({"row": row_number, "reason": "missing email"})
            continue
        try:
            email = normalize_email(raw_email)
        except EmailNotValidError:
            skipped += 1
            errors.append({"row": row_number, "email": raw_email, "reason": "invalid email"})
            continue
        if email in seen:
            skipped += 1
            errors.append({"row": row_number, "email": email, "reason": "duplicate in import"})
            continue
        seen.add(email)

        status_value = str(row.get("consent_status") or ConsentStatus.unknown.value).strip()
        if status_value not in ConsentStatus._value2member_map_:
            status_value = ConsentStatus.unknown.value

        payload = ContactCreate(
            email=email,
            first_name=row.get("first_name"),
            last_name=row.get("last_name"),
            company=row.get("company"),
            phone=row.get("phone"),
            source=source,
            consent_status=ConsentStatus(status_value),
            consent_source=row.get("consent_source"),
            tags=[tag.strip() for tag in str(row.get("tags") or "").split(",") if tag.strip()],
            custom_fields={},
        )
        _, created = upsert_contact(db, user_id, payload)
        if created:
            imported += 1
        else:
            updated += 1

    return ImportResult(imported=imported, updated=updated, skipped=skipped, errors=errors)
